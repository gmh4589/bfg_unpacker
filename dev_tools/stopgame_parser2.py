
from selenium import webdriver
from time import sleep
from datetime import datetime
from random import randint
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

a = 1


table = Workbook()
sheet1 = table.create_sheet("SPAR", 0)

head = ['Название', 'Дата выхода']


def resize(sheet):
    for cols in sheet1.columns:
        long = 0

        for cell in cols:
            cellLen = len(str(cell.value))

            if cellLen > long:
                long = cellLen

            letter = cell.column_letter
            sheet.column_dimensions[letter].width = long + 5


def colorize_cell(cell_text, col, r=1, color='90ee90', bold=True):
    c = sheet1.cell(r, col, cell_text)
    c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    c.font = Font(bold=bold)


while True:
    driver = webdriver.Chrome()
    driver.get(f'https://www.igromania.ru/games/?page={a}')

    names = BeautifulSoup(driver.page_source, 'html.parser').select('.GameCard_title__V4i1j')
    dates = BeautifulSoup(driver.page_source, 'html.parser').select('.GameCard_date__ZEatM')
    platforms = BeautifulSoup(driver.page_source, 'html.parser').select('.GameCard_platforms__UGGQ1')
    all_platforms = BeautifulSoup(driver.page_source, 'html.parser').select('.style_control_title__ElIIp')

    ap = [pl.text for pl in all_platforms]

    driver.close()
    driver.quit()

    for i, d in enumerate(names):

        try:
            platforms_list = [p for p in ap if p in platforms[i].text]

            if 'PS' in platforms_list:

                if 'PS2' or 'PS3' or 'PS4' or 'PSP' or 'PS Vita' in platforms_list:
                    platforms_list.remove('PS')

            if 'Xbox' in platforms_list:

                if 'Xbox One' or 'Xbox 360' or 'Xbox Series' in platforms_list:
                    platforms_list.remove('Xbox')

        except IndexError:
            platforms_list = []

        row = sheet1.max_row + 1
        name = d.text
        date = dates[i].text.replace("Вышла: ", "")
        sheet1.cell(row, 1, name)
        sheet1.cell(row, 2, date)

        for pl in platforms_list:

            if pl not in head:
                head.append(pl)

            j = head.index(pl) + 1
            colorize_cell('X', j, row, color='50aa00', bold=False)

        print(f'{names[i].text}: {dates[i].text.replace("Вышла: ", "")}\tПлатформы: {platforms_list}')

    a += 1
    sleep(randint(1, 10))

    if a == 10:
        break


for cl in head:
    colorize_cell(cl, head.index(cl) + 1)

resize(sheet1)
table.save(f'./stop_game_{datetime.now():%d.%m.%Y}.xlsx')
