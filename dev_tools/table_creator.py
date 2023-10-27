
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


class TableCreator:

    def __init__(self):

        self.table = Workbook()
        self.sheet_name = ''
        self.sheet1 = self.table.create_sheet(self.sheet_name, 0)
        self.head = ['Название', 'Дата выхода']

    def resize(self, sheet):
        for cols in self.sheet1.columns:
            long = 0

            for cell in cols:
                cellLen = len(str(cell.value))

                if cellLen > long:
                    long = cellLen

                letter = cell.column_letter
                sheet.column_dimensions[letter].width = long + 5

    def colorize_cell(self, cell_text, col, r=1, color='90ee90', bold=True):
        c = self.sheet1.cell(r, col, cell_text)
        c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c.font = Font(bold=bold)

    def finalize_table(self, name):

        for i in range(len(self.head)):
            self.sheet1.cell(1, i + 1, self.head[i])

        self.resize(self.sheet1)
        self.sheet1.freeze_panes = 'C2'
        self.table.save(f'./{name}_{datetime.now():%d.%m.%Y_%H%M%S}.xlsx')
